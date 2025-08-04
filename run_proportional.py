import os
import sys
import traci
import numpy as np
import random
import csv
import sumolib
import pandas as pd
import openpyxl
from input.data import link, parking_link, parking_area_list, parking_list
from modules.vehicle_generation import generate_vehicle_flow, generate_background_flow
from modules.parking_allocation_proportional import parking_lot_allocation
from modules.parking_behavior import parking_wizard

if 'SUMO_HOME' not in os.environ:
    sys.exit("please declare environment variable 'SUMO_HOME'")
else:
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sumo_gui = os.path.join(os.environ['SUMO_HOME'], 'bin\\sumo-gui')
    sumo = os.path.join(os.environ['SUMO_HOME'], 'bin\\sumo')
    sys.path.append(tools)

def run_sumo(t, right_count, result):
    # store actual parking zone information
    parking_actual = {}
    # store buffer zone information
    parking_buffer = {}
    for i in traci.parkingarea.getIDList():
        parking_actual[i] = 0
        parking_buffer[i] = 0

    veh_dict = {}

    # get the first zone the vehicle is going to, and generate vehicle
    veh_first_zone = generate_vehicle_flow('input/veh_demand.csv')

    # generate background traffic
    # num is the multiple of curb user
    num = 2
    generate_background_flow('input/veh_demand.csv', num)

    step = 0

   # Record the information of the speed-restricted vehicle
    limit_info = {}
    current_speedlimits = {}
    road_limit_info = {}
    # Record background vehicle data
    background_info = {}

    # run simulation
    while step < t:
        traci.simulationStep()
        # detect the actual vehicle id in the parking zone
        for i in traci.parkingarea.getIDList():
            parking_actual[i] = traci.parkingarea.getVehicleCount(i)
        # update the buffer information: if the vehicle is already parked in the space, the vehicle id is deleted from the buffer
        for key in parking_actual:
            if parking_actual[key] == 1:
                parking_buffer[key] = 0
        # depending on whether there exists vehicle in the illegal parking area, see whether the speed limit should be carried out
        for i in traci.parkingarea.getIDList():
            if '_il' in i and traci.parkingarea.getVehicleCount(i) > 0:
                traci.edge.setMaxSpeed(i.split('_')[0], 1)
            elif '_il' in i and traci.parkingarea.getVehicleCount(i) == 0:
                traci.edge.setMaxSpeed(i.split('_')[0], 13.89)

        # Obtain the ID of the speed-limiting vehicle on the speed-limited roads
        for i in ['a7', 'a8', 'a11', 'a12', 'a13', 'a14', 'a17', 'a18',
          'b7', 'b8', 'b11', 'b12', 'b13', 'b14', 'b17', 'b18',
          'c5', 'c7', 'c8', 'c10']:
            speedlimit = traci.lane.getMaxSpeed(i + '_0')
            current_speedlimits[i] = speedlimit
    
            if speedlimit < 5 and i[0] in ['a', 'b', 'c']:
                for j in traci.edge.getLastStepVehicleIDs(i):
                    if j not in limit_info:
                        veh_type_limit = traci.vehicle.getTypeID(j)
                        limit_info[j] = [{'vehicle_id': j, 'start_time': step, 'end_time': None, 'veh_type': veh_type_limit, 'road_id': i}]
                    else:
                        existing_record = None
                        for record in limit_info[j]:
                            if record['road_id'] == i and record['end_time'] is None:
                                existing_record = record
                                break

                        if existing_record:
                            existing_record['end_time'] = None
                        else:
                            veh_type_limit = traci.vehicle.getTypeID(j)
                            limit_info[j].append({'vehicle_id': j, 'start_time': step, 'end_time': None, 'veh_type': veh_type_limit, 'road_id': i})    
                if i not in road_limit_info:
                    road_limit_info[i] = []
                if not any(record['end_time'] is None for record in road_limit_info[i]):
                    road_limit_info[i].append({'road_id': i, 'start_time': step, 'end_time': None})
            else:
                if i in road_limit_info:
                    for record in road_limit_info[i]:
                        if record['end_time'] is None:
                            record['end_time'] = step
        
        # update end_time 
        for veh_id, records in limit_info.items():
            for record in records:
                road_id = record['road_id']
                if record['end_time'] is None:
                    left_road = veh_id not in traci.edge.getLastStepVehicleIDs(road_id)
                    speed_restored = current_speedlimits.get(road_id, 0) >= 5
                    if left_road or speed_restored:
                        record['end_time'] = step

        # get background vehicle data
        for i in traci.vehicle.getIDList():
            if 'background' in i:
                if i not in background_info:
                    # departure time, arrival time, zone, avg_speed
                    background_info[i] = [step, step, i.split('_')[1], []]
                else:
                    background_info[i][1] = step
                    edge = traci.vehicle.getRoadID(i)
                    if edge[0] in ['a', 'b', 'c']:
                        speed = traci.vehicle.getSpeed(i)
                        background_info[i][-1].append(speed)

        if step < 3600:
            # generates the vehicle_type for each space at the specified time and changes the specified poi color
            if step == 0:
                print('0')
                parking_type_info = {}
                # assign parking spaces and change the color of parking spaces
                parking_lot_allocation(parking_type_info, 1)
            # the simulation starts, controlling the vehicle to park
            parking_wizard(step, veh_dict, link, parking_link, parking_area_list,
                           parking_list, parking_buffer, parking_type_info, parking_actual, veh_first_zone, result)
        elif 3600 <= step < 7200:
            # generates the vehicle_type for each space at the specified time and changes the specified poi color
            if step == 3600:
                print('3600')
                parking_type_info = {}
                # assign parking spaces and change the color of parking spaces
                parking_lot_allocation(parking_type_info, 2)
            # the simulation starts, controlling the vehicle to park
            parking_wizard(step, veh_dict, link, parking_link, parking_area_list,
                           parking_list, parking_buffer, parking_type_info, parking_actual, veh_first_zone, result)
        elif 7200 <= step < 10800:
            # generates the vehicle_type for each space at the specified time and changes the specified poi color
            if step == 7200:
                print('7200')
                parking_type_info = {}
                # assign parking spaces and change the color of parking spaces
                parking_lot_allocation(parking_type_info, 3)

            # the simulation starts, controlling the vehicle to park
            parking_wizard(step, veh_dict, link, parking_link, parking_area_list,
                           parking_list, parking_buffer, parking_type_info, parking_actual, veh_first_zone, result)
        elif 10800 <= step < 14400:
            if step == 10800:
                print('10800')
                parking_type_info = {}
                parking_lot_allocation(parking_type_info, 4)

            parking_wizard(step, veh_dict, link, parking_link, parking_area_list,
                           parking_list, parking_buffer, parking_type_info, parking_actual, veh_first_zone, result)
        else:
            if step == 14400:
                print('14400')
                parking_type_info = {}
                parking_lot_allocation(parking_type_info, 5)
            # simulation starts, controlling the vehicle to park
            parking_wizard(step, veh_dict, link, parking_link, parking_area_list,
                           parking_list, parking_buffer, parking_type_info, parking_actual, veh_first_zone, result)

        step += 1
    traci.close()

    # columns to delete
    columns_to_delete = ['temporary_destination', 'final_cruising_time_illegal','temporary_whether_illegal_parking',
                         'temporary_cruising_time', 'last_road_id', 'whether_went_to_intend_block']
    # header
    fieldnames = ['veh_id'] + [col for col in veh_dict['parking_A_1_PUDO_0'].keys() if col not in columns_to_delete]
    # open the CSV file and write the data
    with open(f'output_proportional/output{right_count}.csv', 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        # write column name
        writer.writeheader()
        # write data line by line
        for key, value in veh_dict.items():
            value['actual_passed_through_zones'] = value['parking_zone_sequence'][0:value['actual_passed_through_zones_num']]
            value['average_cruising_speed'] = sum(value['average_cruising_speed']) / len(value['average_cruising_speed']) if value[
                'average_cruising_speed'] else None
            value['average_cruising_speed_illegal'] = sum(value['average_cruising_speed_illegal']) / len(value['average_cruising_speed_illegal']) if value[
                'average_cruising_speed_illegal'] else None

            if value['arrival_time'] != (t - 1):
                if value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 1 and value['illegal_behavior_state'] == 0:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal']
                elif value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 2 and value['illegal_behavior_state'] == 0:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal'] * 2
                elif value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 3 and value['illegal_behavior_state'] == 0:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal'] * 3
                elif value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 1 and value['illegal_behavior_state'] == 1:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal'] + value['cruising_time_threshold_legal'] / 2
                elif value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 1 and value['illegal_behavior_state'] == 2:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal'] + value['actual_cruising_time_illegal']
                elif value['whether_parking_success'] is False and value['actual_passed_through_zones_num'] == 1 and value['illegal_behavior_state'] == 3:
                    value['actual_cruising_time'] = value['cruising_time_threshold_legal'] + value['cruising_time_threshold_legal'] / 2

            # takes the dictionary key as a column of data
            value['veh_id'] = key
            # deletes the specified column
            for column in columns_to_delete:
                del value[column]
            writer.writerow(value)
    
    # Save the data of speed-limited vehicles
    with open(f'output_proportional/output_limit{right_count}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Vehicle ID', 'Start Time', 'End Time', 'Vehicle Type', 'Road ID'])
        for vehicle_id, records in limit_info.items():
            for record in records:
                writer.writerow([record['vehicle_id'], record['start_time'], record['end_time'], record['veh_type'], record['road_id']])

    # Save the speed limit road data 
    with open(f'output_proportional/output_limited_road{right_count}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Road ID', 'Start Time', 'End Time'])
        for road_id, records in road_limit_info.items():
            for record in records:
                writer.writerow([record['road_id'], record['start_time'], record['end_time']])

    # Save the background traffic information
    for vehicle_id, info in background_info.items():
        speed_list = info[3]
        if speed_list:
            avg_speed = sum(speed_list) / len(speed_list)
        else:
            avg_speed = 0
        info[3] = avg_speed

    with open(f'output_proportional/output_background{right_count}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Vehicle ID', 'Departure Time', 'Arrival Time', 'Area', 'Average Speed'])
        for vehicle_id, info in background_info.items():
            writer.writerow([vehicle_id, info[0], info[1], info[2], info[3]])

if __name__ == '__main__':
     # get vehicle's intend_parking_block data
     # read Excel file
     workbook = openpyxl.load_workbook('input/vehicle_demand_block.xlsx')
     # process three sheets
     result = {}
     for sheet_name in['zone1', 'zone2', 'zone3']:
        worksheet = workbook[sheet_name]
        header =[cell.value for cell in worksheet[1]]

        for row in range(2, worksheet.max_row + 1):
            row_data = [worksheet.cell(row=row, column=col).value for col in range(1, worksheet.max_column + 1)]
            key = (row_data[0], row_data[1], row_data[2])
            if sheet_name == 'zone3':
                 values = [float(x) for x in row_data[3:7] if x is not None]
                 total = sum(values)
                 probabilities = [x / total for x in values]
                 # the data structure of sheet 3 is different from the other 2
                 result[key] = dict(zip(header[3:7], probabilities))
            else:
                values = [float(x) for x in row_data[3:11] if x is not None]
                total = sum(values)
                probabilities =[x / total for x in values]
                result[key] = dict(zip(header[3:11], probabilities))
             
     # simulation time (s)
     t = 1800 * 2
     # the number of simulation successfully complete requirement, and tolerance for unsuccessful running
     num1, num2 = 1, 2
     # number of simulation runs, number of successful running, number of unsuccessful running
     count, right_count, error_count = 0, 0, 0
     for i in range(3):
         print(f'The {count} simulation')
         # create a dictionary to store the data
         parking_allocation_data = {}
         if right_count == num1:
            print(
                f'The simulation has been run for {count} times.\n'
                f'The correct number of simulations is {right_count} times.\n'
                f'The number of simulation errors is {error_count}')
            break
         if error_count == num2:
            print(
                f'There are too many errors in the simulation.\n'
                f'The simulation has been run for {count} times.\n'
                f'The correct number of simulations is {right_count} times.\n'
                f'The number of simulation errors is {error_count}. ')
            break
         try:
             # count simulation times
             count += 1
             sumoCmd = [sumo_gui, "-c", "run.sumocfg"]
             # traci starts the simulation
             traci.start(sumoCmd)
             # run the main function
             run_sumo(t, right_count, result)
             right_count += 1  
         except Exception as e:
             error_count += 1
             traci.close()
             print(f"simulation errors occur after the {count} count:error is: {str(e)}")
             continue
         